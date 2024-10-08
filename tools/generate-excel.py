import os
import sys
import yaml
import jsonref
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill,Font

# reguirements
# openpyxl = "openpyxl==3.1.5"
# jsonref = "jsonref==1.1.0"
# pyyaml = "pyyaml==6.0.2
# 

def export_to_excel(ws, schema, types):

    def format(item, style):
        #if style.get("word_wrap"):
        #    alignment = Alignment(vertical="top", wrap_text=style.get("word_wrap"))
        #else: 
        alignment = None
        font = Font(name=style.get("font"), size=style.get("size"), bold=style.get("bold"), color=style.get("fgcolor"))

        if style.get("bgcolor"):
            fill = PatternFill(start_color=style.get("bgcolor"), end_color=style.get("bgcolor"), fill_type="solid")
        else:
            fill = None

        if not isinstance(item, tuple):
            item = [item]
        for cell in item:
            if alignment:
                cell.alignment = alignment 
            if font:
                cell.font = font
            if fill:
                cell.fill = fill            

    # Append the title and header rows
    ws.append(["PACT Simplified Tech Specs", "", "", "", "", "", "", "", "", "", ""])
    ws.append([schema["info"]["version"], "", "", "", "", "", "", "", "", ""])
    ws.append([
        "Property",                     # Column A
        "Mandatory?",	                # Column B
        "Methodology Attribute Name",   # Column C
        "Link to Methodology",	        # Column D
        "User Friendly Description of Attribute",     # Column E
        "Unit",	                        # Column F
        "Accepted Value(s)",	        # Column G
        "Example 1 (Dummy data)",	    # Column H
        "Example 2 (Dummy data)",	    # Column I
        "Example 3 (Dummy data)"        # Column J
    ])
    # TODO: descriptions
    ws.append(["description", "description", "", "", "", "", "", "", "", ""])

    # Set cell widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 30

    
    fontname = "Aptos Narrow"
    title_style = dict(font = fontname, size = 16, bold = True, bgcolor = "465C66", fgcolor = "FFFFFF")
    subtitle_style = dict(font = fontname, size = 16, bold = False, bgcolor = "465C66", fgcolor = "FFFFFF")
    header_style = dict(font = fontname, bgcolor = "D9D9D9")
    heading_style = dict(font = fontname, bold = True, bgcolor = "09094E", fgcolor = "FFFFFF")
    normal_style = dict(font = fontname)
    bold_style = dict(font = fontname, bold = True)
    wrap_style = dict(font = fontname, word_wrap = True)

    # format the first rows with the title fill
    format(ws[1], title_style)
    format(ws[2], subtitle_style)
    format(ws[3], header_style)
    format(ws[4], header_style)
    format(ws["B4"], normal_style)


    def get_type_description(info):
        type_description = ""
        if info.get("type", "") == "array":
            type_description = "list of " + get_type_description(info["items"])
        if info.get("type", "") == "object":
            type_description = "object"
        type_description += " " + info.get("format", "") + "\n"
        type_description += " " + info.get("comment", "") + "\n"
        type_description += "|".join(info.get("enum", [])) + "\n"
        type_description = type_description.strip()
        if (type_description == ""):
            type_description = info.get("type","")
        elif not info.get("type","") in ["object","array"]:
            type_description += "\n(" + info.get("type","") + ")"
        return type_description

    # Inner function to write a property to the worksheet
    def write_property(name, info, parent, level):
        # Extract the type and description of the property
        type = info.get("type", "")
        
        if type == "object":
            write_type(name, info, level + 1)
            return
        
        type_description = get_type_description(info)
        description = info.get("description", "N/A")
        examples = info.get("examples", []) + ['','','']
        mandatory = name in parent.get("required", [])
        
        # Append a row to the worksheet
        ws.append([
            name, 
            "M" if mandatory else "O",
            info.get("title", name),
            "-",
            description.rstrip(), # remove last newline from the description
            "-",
            type_description, 
            str(examples[0]),
            str(examples[1]),
            str(examples[2])
            ])
        # Indent the first cell of the row just added
        format(ws[ws.max_row], normal_style)
        format(ws[ws.max_row][0], bold_style)
        ws[ws.max_row][0].alignment = Alignment(indent=level)

        if info.get("type", None) == "array" and info["items"].get("type") == "object":
            write_type(None, info["items"], level + 1)


        
    # Inner function to write a type to the worksheet
    def write_type(name, info, level=0):
        if info.get("title"):
            # Append a row for the type itself and set background color to blue
            ws.append([name + ": " + info["title"], "", "", "", "", "", "", "", "", ""])
            format(ws[ws.max_row], heading_style)

        for prop_name, prop_info in info.get("properties", {}).items():
            # Extract the type and description of the property
            write_property(prop_name, prop_info, info, level)


    # Find the specified types in the schema
    for name in types:
        type = schema["components"]["schemas"][name]
        write_type(name, type)


    # Set word wrap for all cells in the description colum
    for cell in ws["A"] :
        cell.alignment = Alignment(vertical="top", indent=cell.alignment.indent)
    for cell in ws["B"] + ws["C"] + ws["D"] + ws["E"] + ws["F"] + ws["G"] + ws["H"] + ws["I"] + ws["J"]:
        cell.alignment = Alignment(vertical="top")
    for cell in ws["E"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws["G"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# check if not debugging:
if False:
    # input_path = "pact-openapi-2.2.1-wip.yaml"
    input_path = "pact-openapi-2.2.0.yaml"
else:
    if len(sys.argv) < 2:
        print("Usage: python3 generate-excel.py <input-path>")
        print("This script generates an Excel file from a OpenAPI schema.")
        print("")
        print("Example:")
        print("python3 generate-excel pact-openapi-2.2.1-wip.yaml")
        print()
        exit()
    input_path = sys.argv[1]

    
# Load the schema from the file
with open(input_path) as file:
    schema1 = yaml.safe_load(file)
schema = jsonref.replace_refs(schema1, merge_props=True)

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "PACT Simplified Data Model"
ws.sheet_view.zoomScale = 140

export_to_excel(ws, schema, ["ProductFootprint"])

# Save the workbook to a file

# change extension to .xlsx using python path manipulation

output_path = os.path.basename(input_path)
output_path = output_path.replace('-openapi-', '-simplified-model-')
output_path = output_path.replace(".yaml", "") + ".xlsx"
output_path = os.path.join(os.path.dirname(input_path), output_path)
print(output_path)
wb.save(output_path)
